'''
Created on Dec 18, 2018

@author: rajat.arora07
'''

import openpyxl
import os


class ReadExcel():
    '''
        The class provides an OOP solution to read an excel file 
        and store its data in a list.
    '''

    def __init__(self, path):
        '''
            Sets the path and initiates the user_list.
        '''
        self.__path = path
        self.__user_list = list()
    
    # Setters and Getters for Path and User list respectively.
    def set_path(self, path):
        self.__path = path
    
    def get_path(self):
        return self.__path
    
    def set_user_list(self, data):
        self.__user_list.append(data)
        
    def get_user_list(self):
        return self.__user_list
    
    # Function to read excel.
    def read_excel(self):
        
        # workbook object is created 
        wb_obj = openpyxl.load_workbook(self.get_path()) 
        
        sheet_obj = wb_obj.active 
        # Get the maximum no. of rows and columns.
        m_row = sheet_obj.max_row 
        m_col = sheet_obj.max_column
                
        # Looping from 2nd row to fetch all the data.
        # Ist row is excluded as it would consist of column names.
        for i in range(2,m_row+1): 
            # Dictionary for a single user data.
            user_data = dict()  
            data = dict()
            for j in range(1,m_col+1):
                cell_obj = sheet_obj.cell(row = i, column = j) 
                data[sheet_obj.cell(row = 1, column = j).value] =  cell_obj.value
            
            # User data dictionary will have Date column as key and 
            # the whole user data as its value.
            user_data[data['Date']] = data
            self.set_user_list(user_data)
        
# Path to file.
path = "D:\Eclipse_Codes\Pypy\geeks_for_geeks\Read_excel.xlsx" 
if os.path.isfile(path):
    ob = ReadExcel(path)
    ob.read_excel()
    print ob.get_user_list()
else:
    print 'No file/ Directory'